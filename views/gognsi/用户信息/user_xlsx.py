#处理xlsx


from openpyxl import load_workbook
import  config
import  json

def xlsx(sheet):
    # print('sheet名称=', sheet)
    # print('根据坐标获取=',sheet['A1'].value,sheet['B1'].value)
    # print('根据坐标获取=', sheet['A2'].value, sheet['B2'].value)

    rows = sheet.rows
    # print(rows)
    title = []
    listData=[]
    index = 0
    # 迭代所有的行
    for row in rows:
        # print(row)
        # 获取整行数据，返回list
        line = [col.value for col in row]
        # print(line)

        #第一行是标题
        if 0 == index:
            i = 0
            for i in range(len(line)):
                title.append(line[i])
            # print('title=',title)
        else:
            i = 0
            map={}
            for i in range(len(line)):
                #说明读取完毕，下面的都是空格子
                if title[i]=='user_name' and None==line[i]:
                    return listData
                else:
                    str_title = title[i]
                    map[str_title] = line[i]

                # print(str_title,line[i])
            print('map---------------------------------------------')
            print('map=',map)
            listData.append(map)

        index = index + 1
    # print('------------------listData---------------------------')
    # print('listData=',listData)
    # print(listData)
    # data=json.dumps(listData,ensure_ascii=False)
    # print(data)
    # path=config.base_dirs+'\\其他\\基础评估\\omo_stat_groupname.json'
    # with open(path, 'w', encoding="utf-8") as f:
    #     f.write(data)
    # print('组=',data)
    print('---------------------------------------------')
    return listData


#处理解析后的数据
def dispose(listData):
    for item in listData:

        item['birthday2'] = 0
        if None!=item['birthday']:
            birthday = str(item['birthday'])
            birthday=birthday.split(' ')[0].replace('-','').replace('/','')
            if ''!=birthday:
                item['birthday2']=int(birthday)

        print('item=---------------------------------------------')
        print('itme=',item)
        # print('listData=',listData)
    return  listData


def mian():
    'C:\pythonWorkspace\python_tronado_web\views\gognsi\用户信息'
    path = config.base_dirs + '\\views\gognsi\\用户信息\\秦皇岛-人员编号-处理.xlsx'
    print(path)
    # 加载excel文件
    workbook = load_workbook(path)
    # 获取所有sheet的名称
    sheets = workbook.worksheets
    print('sheet文件=', sheets)
    data=dispose(xlsx(sheets[0]))
    print('len=',len(data))
    return data


if __name__ == '__main__':
    mian()
    # path=config.base_dirs+'\\其他\\用户信息\\处理.xlsx'
    # print(path)
    # # 加载excel文件
    # workbook = load_workbook(path)
    # # 获取所有sheet的名称
    # sheets = workbook.worksheets
    # print('sheet文件=',sheets)
    # dispose(xlsx(sheets[0]))
    # # for sheet in sheets:
    # #     xlsx(sheet)










