
from openpyxl import load_workbook

from openpyxl import load_workbook

#相对路径，找到需要打开的文件位置
workbook = load_workbook(r'C:\test\1.xlsx')
#获取当前活跃的sheet,默认是第一个sheet
booksheet = workbook.active


#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns

i = 0
# 迭代所有的行
for row in rows:
    # print(row)
    #获取整行数据，返回list
    line = [col.value for col in row]
    if 0==i:
        print('title=',line)
    else:
        print('value=', line)
        v = [col for col in line]
        print(v)
    i = i + 1




